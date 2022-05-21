# merging the CSV and xlsx files

import calendar
import os
import time
import tkinter as tk
from datetime import datetime, timedelta
from tkinter import *
from tkinter import ttk, filedialog

import openpyxl as ex
import pandas as pd
from tkcalendar import DateEntry

root = tk.Tk()

root.geometry("400x300")
root.title('Santos Productivity')
Label(root, text="Click the Button to browse the files", font='Georgia 13').pack(pady=25)

root.file_name = ''

startdate = DateEntry(root, locale='en_US', date_pattern='MM/dd/yyyy')
startdate.place(x=40, y=140)

enddate = DateEntry(root, locale='en_US', date_pattern='MM/dd/yyyy')
enddate.place(x=210, y=140)


def open_file():
    root.file_name = filedialog.askopenfilename(parent=root, initialdir="/")


def timestamp_cov(date):
    try:
        datetime.strptime(date, "%m/%d/%Y %I:%M %p")
        data = "%m/%d/%Y %I:%M %p"
    except:
        data = "%m/%d/%y %H:%M"

    return calendar.timegm(time.strptime(date, data))


def ts_2_hrd(date):
    # return datetime.utcfromtimestamp(date)
    return datetime.fromtimestamp(date).strftime('%Y-%m-%d')


def main():
    df = pd.read_csv(root.file_name)

    df = df[df['Tested On'].notna()]

    df['Timestamp'] = df.apply(lambda row: timestamp_cov(row['Tested On']), axis=1)

    df['Human read datetime'] = df.apply(lambda row: ts_2_hrd(row['Timestamp']), axis=1)

    start_date = str(startdate.get_date())

    e_date = (enddate.get_date())
    td = timedelta(1)
    end_date = str(e_date + td)

    filtered_df = df.loc[(df['Human read datetime'] >= start_date) & (df['Human read datetime'] < end_date)]

    sort = pd.DataFrame(filtered_df)
    sort.pop('Human read datetime')
    sort.pop('Timestamp')
    filtered_df.to_excel("Human read.xlsx")
    print(sort.groupby(['Tested By']).count())
    # dig.to_excel('checking.xlsx')

    sum_of_count = (sort.groupby(['Tested By']).count()).sum()
    print(sum_of_count)


b_chooseFile = ttk.Button(root, text="Chose File", width=20, command=open_file)
b_chooseFile.place(x=90, y=100)

b_chooseFile = ttk.Button(root, text="Run", width=10, command=main)
b_chooseFile.place(x=130, y=220)

root.mainloop()


path = "/Users/kamuthuh/PycharmProjects/Projectsantos-Productivity_tracker/temp.xlsx"

wb = ex.load_workbook(path)

os.remove("/Users/kamuthuh/PycharmProjects/Projectsantos-Productivity_tracker/temp.xlsx")
print("File Removed!")

sheet = wb.active

cell_obj = sheet.cell(row=1, column=1)

print(sheet.max_row)
max_row = sheet.max_row

res, file_ = {}, []

# Loop will print all rows name
for i in range(2, max_row + 1):
    cell_obj_1 = sheet.cell(row=i, column=2)
    cell_obj_2 = sheet.cell(row=i, column=5)

    file_.append((cell_obj_1.value, cell_obj_2.value))

print(f" length of file: {len(file_)}")


def group_excel_by_date(file: list):
    """
    file = [('abcd', '18_05_2022'), {'abcd', '17_05_2022')]
    {
	"abcd": {
		"18_05_2022": 31,
		"17_05_2022": 28,
		"15_05_2022": 0
	}
}
    :return: dict
    """

    list_rows = file

    for item in list_rows:
        # if the name hasn't been entered to dict yet
        if item[0] not in res:
            res[item[0]] = {item[1]: 1}
        elif item[0] in res:
            if item[1] in res[item[0]]:
                res[item[0]][item[1]] = res[item[0]][item[1]] + 1
            else:
                res[item[0]][item[1]] = 1

    return res


result = group_excel_by_date(file_)
print(result)
