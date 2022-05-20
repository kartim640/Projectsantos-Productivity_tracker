import os
import openpyxl as ex
from collections import OrderedDict


path = "/Users/kamuthuh/PycharmProjects/Projectsantos-Productivity_tracker/Human read.xlsx"

wb = ex.load_workbook(path)


sheet = wb.active

cell_obj = sheet.cell(row=1, column=1)

print(sheet.max_row)
max_row = sheet.max_row

res, file_ = {}, []

# Loop will print all rows name
for i in range(2, max_row + 1):
    cell_obj_1 = sheet.cell(row=i, column=2)
    cell_obj_2 = sheet.cell(row=i, column=5)

    file_.append((cell_obj_1.value, cell_obj_2.value))

print(f" length of file: {len(file_)}")


def group_excel_by_date(file: list):
    """
    file = [('abcd', '18_05_2022'), {'abcd', '17_05_2022')]
    {
	"abcd": {
		"18_05_2022": 31,
		"17_05_2022": 28,
		"15_05_2022": 0
	}
}
    :return: dict
    """

    list_rows = file

    for item in list_rows:
        # if the name hasn't been entered to dict yet
        if item[0] not in res:
            res[item[0]] = {item[1]: 1}
        elif item[0] in res:
            if item[1] in res[item[0]]:
                res[item[0]][item[1]] = res[item[0]][item[1]] + 1
            else:
                res[item[0]][item[1]] = 1

    return res


result = group_excel_by_date(file_)
print(result)
